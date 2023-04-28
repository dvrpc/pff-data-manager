import os
import psycopg2 as psql
import openpyxl
import pandas as pd
import requests
import json
import csv
from shapely.geometry import shape
from datetime import datetime
from utils.settings import FREIGHTDB_CONN, CENSUS_API_KEY

con = psql.connect(FREIGHTDB_CONN)
cur = con.cursor()

#loops through *.xlsx files in dir and inserts their contents into port_tonnage table
def load_tonnage(dir, overwrite=False):
    try:
        files = os.listdir(dir)
    except FileNotFoundError:
        print("directory", dir, "was not found")
    else:
        for file in files:
            if file.endswith(".xlsx"):
                try:
                    #since the sheet names are inconsistent, the getPortSheetName function is used to find a valid sheet name for the specified file
                    sheetName = getPortSheetName(os.path.join(dir, file))
                    #gets the year listed at the top of the specified sheet
                    year = getYear(os.path.join(dir, file), sheetName)
                except:
                    print("Could not determine year for", file, "in", dir)
                else:
                    #check if data for the year already exists on the server
                    if not yearExists(year) or overwrite:
                        print("Loading Tonnage for", year)
                        try:
                            ports = pd.read_excel(os.path.join(dir, file), sheet_name=sheetName, header=4)
                            for row in ports.itertuples(name=None):
                                #sanitizes port names and inserts data into port_tonnage table
                                cur.execute(f"""
                                INSERT INTO army_corps.port_tonnage(port_name, year, total_tons, domestic_tons, foreign_tons, import_tons, export_tons)
                                VALUES ('{row[getPortSheetColumnNumber(ports, "port_name")].replace("'","")}', {year}, {row[getPortSheetColumnNumber(ports, "total")]}, {row[getPortSheetColumnNumber(ports, "domestic")]}, {row[getPortSheetColumnNumber(ports, "foreign")]}, {row[getPortSheetColumnNumber(ports, "import")]}, {row[getPortSheetColumnNumber(ports, "export")]})
                                """)
                        except Exception as e:
                            print("Failed to load data for", year, "from", os.path.join(dir, file))
                            print(e)
                    else:
                        print("Tonnage for", year, "already exists")
            else:
                continue
        #commits changes
        con.commit()

def getPortSheetColumnNumber(portsFrame, colName):
    if colName == "port_name":
        for possibleName in ["PORT_NAME"]:
            if possibleName in portsFrame.columns.to_list():
                return portsFrame.columns.get_loc(possibleName) + 1
        return None
    elif colName == "total":
        for possibleName in ["TOTAL","GRAND_TOTAL"]:
            if possibleName in portsFrame.columns.to_list():
                return portsFrame.columns.get_loc(possibleName) + 1
        return None
    elif colName == "domestic":
        for possibleName in ["DOMESTIC"]:
            if possibleName in portsFrame.columns.to_list():
                return portsFrame.columns.get_loc(possibleName) + 1
        return None
    elif colName == "foreign":
        for possibleName in ["FOREIGN","FOREIGN_TOTAL"]:
            if possibleName in portsFrame.columns.to_list():
                return portsFrame.columns.get_loc(possibleName) + 1
        return None
    elif colName == "import":
        for possibleName in ["IMPORTS"]:
            if possibleName in portsFrame.columns.to_list():
                return portsFrame.columns.get_loc(possibleName) + 1
        return None
    elif colName == "export":
        print(portsFrame.columns.isin(["EXPORTS"])[0])
        for possibleName in ["EXPORTS"]:
            if possibleName in portsFrame.columns.to_list():
                return portsFrame.columns.get_loc(possibleName) + 1
        return None
    else:
        return None

#retrieves the port_codes and port_names for all principal ports from a Bureau of Transportation Statistics REST service and inserts them into the army_corps.principal_ports table 
"""
def load_principal_ports():
    try:
        #set the year based on the layer's last edited date
        layerUrl = "https://services7.arcgis.com/n1YM8pTrFmm7L4hs/ArcGIS/rest/services/ndc/FeatureServer/1?f=json"
        lastEditDate = requests.get(layerUrl).json()["editingInfo"]["lastEditDate"]
        year = datetime.fromtimestamp(float(lastEditDate)/1000).year
        print("Loading principal ports")
        
        #gets and iterates through the features from the REST service
        url = "https://services7.arcgis.com/n1YM8pTrFmm7L4hs/ArcGIS/rest/services/ndc/FeatureServer/1/query?where=1%3D1&outFields=PORT_NAME,PORT,TYPE&returnGeometry=true&outSR=4326&f=geojson"
        features = requests.get(url).json()["features"]
        for f in features:
            try:
                #ensures port_code has the correct number of leading zeros, sanitizes port_name, and inserts data into principal_ports table
                row = f["properties"]
                port_name=row["PORT_NAME"].replace("'","")
                port_type=row["TYPE"]
                port_code = port_type + str(row["PORT"]).zfill(4)
                wkt = shape(f["geometry"]).wkt
                cur.execute(f"INSERT INTO army_corps.principal_ports(port_code, port_name, port_type, year, geom) VALUES ('{port_code}', '{port_name}', '{port_type}', {year}, '{wkt}')")
            except Exception as e:
                print("Failed to insert", row)
                print(e)
        con.commit()
    except:
        print("Failed to load principal ports")
"""

#loops through possible names and returns the first name that matches a sheet in the specified *xlsx file
def getPortSheetName(path):
    possibleNames = ["Ports_by_Name", "Port_Name"]
    xlsx = openpyxl.load_workbook(path)
    sheets = xlsx.sheetnames
    for name in possibleNames:
        if name in sheets:
            return name
    raise Exception("getPortSheetName was unable to determine the proper sheet name for the workbook {wb}".format(wb=path))

#looks in the specified sheet to find the year
def getYear(path, sheetName):
    try:
        sheet = openpyxl.load_workbook(path)[sheetName]
        year = sheet["B2"].value
        year = year[year.index("in") + 3:]
        return year
    except:
        raise Exception("Error in getYear()")

#checks if data for a year already exists in the port_tonnage table, returns a boolean 
def yearExists(year):
    cur.execute("SELECT EXISTS(SELECT 1 FROM army_corps.port_tonnage WHERE year = %s)", (year,))
    val = cur.fetchone()
    if val[0] == True:
        return True
    else:
        return False

#geojson importer function, currently unused but may be useful in the future
"""
def insert_port_geojson(path, year):
    file = open(path, "r")
    data = json.load(file)
    features = data["features"]
    for f in features:
        row = f["properties"]
        port_code = str(row["PORT"]).zfill(4)
        port_name=row["PORT_NAME"].replace("'","")
        wkt = shape(f["geometry"]).wkt
        cur.execute(f"INSERT INTO army_corps.principal_ports(port_code, port_name, year, geom) VALUES ('{port_code}', '{port_name}', {year}, '{wkt}')")
    con.commit()
"""

#inserts an arbitrary csv file into an exisitng sql table given the path to the csv and the sql table name (including schema name)
#TABLE AND COLUMNS MUST ALREADY EXIST 
#FIRST ROW OF CSV MUST BE A HEADER THAT MATCHES COLUMN NAMES EXACTLY
def insert_csv(path, table_name):
    try:
        file = open(path, "r", encoding='utf-8-sig')
        data = csv.reader(file)
        line = 0 
        col_names = ""
        for row in data:
            if line == 0:
                col_names = ", ".join(row)
                print(col_names)
                line = line + 1
            else:
                values = []
                for i in range(len(row)):
                    if row[i].isdigit():
                        #row[i] is int
                        values.append(row[i])
                    elif row[i].replace(".","").replace("-","").isdigit():
                        #row[i] is float
                        values.append(row[i])
                    elif row[i] == None or row[i] == "" or row[i] == "NULL":
                        #row[i] is null
                        values.append("NULL")
                    else:
                        #row[i] is str
                        values.append("'" + row[i] +  "'")
                values = ", ".join(values)
                cur.execute(f"INSERT INTO {table_name}({col_names}) VALUES ({values})")
                line = line + 1
        con.commit()            
    except FileNotFoundError:
        print("File not found at", path)
    except Exception as e:
        print("insert_csv failed to insert", path, "into table", table_name)
        print(e)
        con.rollback()
    